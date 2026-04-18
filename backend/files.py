"""File handling: upload to Supabase Storage, text extraction, cleanup."""

import os
import io
import base64
import uuid
import httpx
from typing import Optional, Dict, Any, List, Tuple

SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
BUCKET_NAME = "attachments"

# File type categories
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
PDF_EXTENSIONS = {".pdf"}
DOCX_EXTENSIONS = {".docx"}
XLSX_EXTENSIONS = {".xlsx"}
TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".xml", ".html", ".py", ".js", ".ts", ".jsx", ".tsx", ".css", ".sql", ".sh", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".log"}

ALL_SUPPORTED = IMAGE_EXTENSIONS | PDF_EXTENSIONS | DOCX_EXTENSIONS | XLSX_EXTENSIONS | TEXT_EXTENSIONS


def _storage_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }


def ensure_bucket():
    """Create the attachments bucket if it doesn't exist."""
    with httpx.Client() as client:
        # Check if bucket exists
        r = client.get(
            f"{SUPABASE_URL}/storage/v1/bucket/{BUCKET_NAME}",
            headers=_storage_headers(),
        )
        if r.status_code == 200:
            return

        # Create bucket
        client.post(
            f"{SUPABASE_URL}/storage/v1/bucket",
            headers={**_storage_headers(), "Content-Type": "application/json"},
            json={
                "id": BUCKET_NAME,
                "name": BUCKET_NAME,
                "public": False,
            },
        )


def get_file_category(filename: str) -> str:
    """Return the category of a file based on extension."""
    ext = os.path.splitext(filename)[1].lower()
    if ext in IMAGE_EXTENSIONS:
        return "image"
    if ext in PDF_EXTENSIONS:
        return "pdf"
    if ext in DOCX_EXTENSIONS:
        return "docx"
    if ext in XLSX_EXTENSIONS:
        return "xlsx"
    if ext in TEXT_EXTENSIONS:
        return "text"
    return "unknown"


def upload_file(session_id: str, filename: str, content: bytes, content_type: str) -> Dict[str, Any]:
    """Upload a file to Supabase Storage under session_id/."""
    ext = os.path.splitext(filename)[1].lower()
    storage_filename = f"{uuid.uuid4().hex}{ext}"
    storage_path = f"{session_id}/{storage_filename}"

    with httpx.Client() as client:
        r = client.post(
            f"{SUPABASE_URL}/storage/v1/object/{BUCKET_NAME}/{storage_path}",
            headers={
                **_storage_headers(),
                "Content-Type": content_type,
            },
            content=content,
        )
        r.raise_for_status()

    category = get_file_category(filename)

    return {
        "storage_path": storage_path,
        "original_name": filename,
        "content_type": content_type,
        "size": len(content),
        "category": category,
    }


def delete_session_files(session_id: str):
    """Delete all files for a session from Supabase Storage."""
    with httpx.Client() as client:
        # List files in session folder
        r = client.post(
            f"{SUPABASE_URL}/storage/v1/object/list/{BUCKET_NAME}",
            headers={**_storage_headers(), "Content-Type": "application/json"},
            json={"prefix": f"{session_id}/", "limit": 1000},
        )
        if r.status_code != 200:
            return

        files = r.json()
        if not files:
            return

        # Delete all files
        paths = [f"{session_id}/{f['name']}" for f in files]
        client.delete(
            f"{SUPABASE_URL}/storage/v1/object/{BUCKET_NAME}",
            headers={**_storage_headers(), "Content-Type": "application/json"},
            json={"prefixes": paths},
        )


def get_file_url(storage_path: str) -> str:
    """Get a signed URL for a file (valid for 1 hour)."""
    with httpx.Client() as client:
        r = client.post(
            f"{SUPABASE_URL}/storage/v1/object/sign/{BUCKET_NAME}/{storage_path}",
            headers={**_storage_headers(), "Content-Type": "application/json"},
            json={"expiresIn": 3600},
        )
        r.raise_for_status()
        data = r.json()
        return f"{SUPABASE_URL}/storage/v1{data['signedURL']}"


def download_file(storage_path: str) -> bytes:
    """Download a file from Supabase Storage."""
    with httpx.Client() as client:
        r = client.get(
            f"{SUPABASE_URL}/storage/v1/object/{BUCKET_NAME}/{storage_path}",
            headers=_storage_headers(),
        )
        r.raise_for_status()
        return r.content


def extract_text(filename: str, content: bytes) -> Optional[str]:
    """Extract text content from a file for inclusion in prompts."""
    category = get_file_category(filename)

    if category == "text":
        try:
            return content.decode("utf-8", errors="replace")
        except Exception:
            return None

    if category == "pdf":
        try:
            import pymupdf
            doc = pymupdf.open(stream=content, filetype="pdf")
            text_parts = []
            for page in doc:
                text_parts.append(page.get_text())
            doc.close()
            return "\n".join(text_parts)
        except Exception as e:
            print(f"PDF extraction error: {e}")
            return None

    if category == "docx":
        try:
            from docx import Document
            doc = Document(io.BytesIO(content))
            return "\n".join(p.text for p in doc.paragraphs)
        except Exception as e:
            print(f"DOCX extraction error: {e}")
            return None

    if category == "xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True)
            text_parts = []
            for sheet in wb.worksheets:
                text_parts.append(f"## Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    text_parts.append("\t".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(text_parts)
        except Exception as e:
            print(f"XLSX extraction error: {e}")
            return None

    return None


def build_message_content(
    text: str,
    attachments: List[Dict[str, Any]],
) -> Any:
    """
    Build the message content for OpenRouter API.

    For text-only messages, returns a plain string.
    For messages with images, returns a list of content parts (vision format).
    Text from documents is prepended to the user's text.
    """
    doc_texts = []
    image_parts = []

    for attachment in attachments:
        category = attachment["category"]
        content_bytes = download_file(attachment["storage_path"])

        if category == "image":
            # Convert to base64 data URL for vision models
            b64 = base64.b64encode(content_bytes).decode("utf-8")
            mime = attachment.get("content_type", "image/png")
            image_parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}"},
            })
        else:
            # Extract text and prepend
            extracted = extract_text(attachment["original_name"], content_bytes)
            if extracted:
                name = attachment["original_name"]
                doc_texts.append(f"[Attached file: {name}]\n{extracted}")

    # Build the text portion
    full_text = text
    if doc_texts:
        docs_section = "\n\n".join(doc_texts)
        full_text = f"{docs_section}\n\n---\n\nUser question: {text}"

    # If no images, return plain string (cheaper, compatible with all models)
    if not image_parts:
        return full_text

    # Vision format: list of content parts
    parts = [{"type": "text", "text": full_text}]
    parts.extend(image_parts)
    return parts
